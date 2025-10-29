import streamlit as st
import pandas as pd
import re
from datetime import datetime, timedelta
import io
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import os
import paramiko
from io import StringIO

@st.cache_data
def parse_log_content(content):
    data = []
    
    # Dividir o conteúdo em seções de arquivo
    file_sections = re.split(r'={70,}', content)
    
    for section in file_sections:
        if not section.strip():
            continue
            
        file_data = parse_file_section(section)
        if file_data:
            data.extend(file_data)
    
    return data

def parse_file_section(section):
    lines = section.strip().split('\n')
    
    # Extrair informações básicas do arquivo
    rcs_file = None
    working_file = None
    revisions = []
    current_revision = None
    in_message = False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        if line.startswith("RCS file:"):
            rcs_file = line.split(":", 1)[1].strip()
        elif line.startswith("Working file:"):
            working_file = line.split(":", 1)[1].strip()
        elif line.startswith("revision "):
            # Se já estávamos processando uma revisão, finalizá-la
            if current_revision:
                # Processar a mensagem coletada
                current_revision['message'] = clean_message(current_revision['message'])
                revisions.append(current_revision)
            
            # Iniciar nova revisão
            revision_match = re.match(r'revision\s+([\d.]+)', line)
            if revision_match:
                current_revision = {
                    'revision': revision_match.group(1),
                    'date': None,
                    'author': None,
                    'message': []
                }
                in_message = False
        elif line.startswith("date:") and current_revision:
            # Extrair data e autor
            date_match = re.search(r'date:\s*([^;]+);', line)
            author_match = re.search(r'author:\s*([^;]+);', line)
            
            if date_match:
                current_revision['date'] = date_match.group(1).strip()
            if author_match:
                current_revision['author'] = author_match.group(1).strip()
        elif current_revision and (line.startswith("#") or in_message or (line and not line.startswith("branches:") and not line.startswith("===="))):
            # Filtrar mensagem do commit
            if not in_message and line and not line.startswith("----------------------------"):
                in_message = True
            
            if in_message and line and not line.startswith("----------------------------"):
                current_revision['message'].append(line)
    
    # Adicionar a última revisão
    if current_revision:
        current_revision['message'] = clean_message(current_revision['message'])
        revisions.append(current_revision)
    
    if rcs_file and working_file and revisions:
        # Para cada revisão, criar uma entrada no DataFrame
        result = []
        for rev in revisions:
            # Extrair informações PDR da mensagem
            pdr_info = extract_pdr_info(rev['message'])
            
            # Extrair nome do arquivo do caminho (última parte após /)
            file_name = extract_filename_from_path(rcs_file)
            
            # Extrair Centro e Estado do caminho
            centro, estado = extract_centro_estado(rcs_file)
            
            # Separar data e hora
            data_str, hora_str = parse_date_time(rev['date'])
            
            result.append({
                'rcs_file': clean_path(rcs_file),
                'working_file': file_name,
                'revision': rev['revision'],
                'author': rev['author'],
                'date': data_str,
                'time': hora_str,
                'message': rev['message'],
                'is_pdr': rev['message'].startswith('#') if rev['message'] else False,
                'pdr_classification': pdr_info['classification'],
                'pdr_time': pdr_info['time_minutes'],
                'pdr_description': pdr_info['description'],
                'centro': centro,
                'estado': estado
            })
        return result
    
    return []

def extract_centro_estado(rcs_file):
    """Extrai Centro e Estado do caminho do arquivo"""
    centro = None
    estado = "GERAL"
    
    # Padrão para encontrar /telas/Centro/...
    pattern = r'/telas/Centro/([^/]+)(?:/([^/]+)(?:/|$))?'
    match = re.search(pattern, rcs_file)
    
    if match:
        centro = match.group(1)
        # Se houver um segundo grupo (estado) E houver mais diretórios após o estado, usar ele
        # Caso contrário, manter "GERAL"
        if match.group(2):
            # Verificar se há mais diretórios após o estado
            # Se o que vem após o centro for imediatamente o nome do arquivo, então é GERAL
            path_after_centro = rcs_file.split(f"/Centro/{centro}/")[-1]
            path_parts = path_after_centro.split('/')
            
            # Se houver mais de uma parte (diretórios adicionais), então o primeiro é o estado
            # Se só tiver uma parte (apenas o nome do arquivo), então é GERAL
            if len(path_parts) > 1:
                estado = match.group(2)
            else:
                estado = "GERAL"
    
    return centro, estado

def parse_date_time(date_str):
    """Separa data e hora, convertendo data para DD/MM/AAAA"""
    if not date_str:
        return None, None
    
    try:
        # Tentar diferentes formatos de data
        for fmt in ['%Y/%m/%d %H:%M:%S', '%Y/%m/%d']:
            try:
                dt = datetime.strptime(date_str, fmt)
                data_formatada = dt.strftime('%d/%m/%Y')
                hora_formatada = dt.strftime('%H:%M:%S') if fmt == '%Y/%m/%d %H:%M:%S' else "00:00:00"
                return data_formatada, hora_formatada
            except ValueError:
                continue
    except:
        pass
    
    return None, None

def extract_filename_from_path(path):
    """Extrai o nome do arquivo do caminho (última parte após /)"""    
    clean_path = path
    if clean_path.endswith(',v'):
        clean_path = clean_path[:-2]    
    parts = clean_path.split('/')
    return parts[-1] if parts else clean_path

def extract_pdr_info(message):
    """Extrai informações PDR da mensagem no formato #Classificacao#Tempo#Descricao"""
    if not message or not message.startswith('#'):
        return {'classification': None, 'time_minutes': None, 'description': None}
    
    # Padrão: #CLASSIFICACAO#TEMPO#DESCRICAO
    pattern = r'^#([^#]+)#([^#]*)#(.+)$'
    match = re.match(pattern, message)
    
    if match:
        classification = match.group(1)
        time_str = match.group(2)
        description = match.group(3)
        
        # Converter tempo para número, se possível
        try:
            time_minutes = float(time_str) if time_str else None
        except ValueError:
            time_minutes = None
            
        return {
            'classification': classification,
            'time_minutes': time_minutes,
            'description': description
        }
    
    return {'classification': None, 'time_minutes': None, 'description': None}

def clean_message(message_lines):
    # Remover linhas vazias no início e no fim
    while message_lines and not message_lines[0].strip():
        message_lines.pop(0)
    while message_lines and not message_lines[-1].strip():
        message_lines.pop()
    
    # Juntar as linhas
    message = ' '.join(message_lines).strip()
    
    # Remover "*** empty log message ***" se presente
    if message == "*** empty log message ***":
        message = ""
        
    return message

def clean_path(path):
    # Remover o prefixo "/export/cvs" se existir
    prefixes = ["/export/cvs", "/export/cvs/"]
    for prefix in prefixes:
        if path.startswith(prefix):
            path = path[len(prefix):]
    
    # Remover ",v" do final se existir
    if path.endswith(',v'):
        path = path[:-2]
    
    return path

def create_excel_file(df):
    """Cria um arquivo Excel a partir do DataFrame"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Log de Telas"
    
    # Adicionar cabeçalhos
    for col_num, column_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_name)
    
    # Adicionar dados
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def normalize_classification(classification, mapping):
    """Normaliza a classificação usando o mapeamento fornecido"""
    if not classification:
        return classification
    
    # Verifica se a classificação está no mapeamento
    for original, target in mapping.items():
        if classification == original:
            return target
    
    return classification

def update_message_classification(message, old_classification, new_classification):
    """Atualiza a classificação na mensagem PDR"""
    if not message or not message.startswith('#') or not old_classification or not new_classification:
        return message
    
    # Substituir a classificação antiga pela nova na mensagem
    # Padrão: #CLASSIFICACAO_ANTIGA#TEMPO#DESCRICAO -> #CLASSIFICACAO_NOVA#TEMPO#DESCRICAO
    pattern = f"^#{re.escape(old_classification)}#"
    replacement = f"#{new_classification}#"
    
    updated_message = re.sub(pattern, replacement, message)
    return updated_message

def apply_classification_mapping_to_dataframe(df, mapping):
    """Aplica o mapeamento de classificações ao DataFrame, atualizando as mensagens"""
    if not mapping or df.empty:
        return df
    
    df_mapped = df.copy()
    
    # Aplicar o mapeamento às mensagens PDR
    for old_classification, new_classification in mapping.items():
        # Encontrar registros com a classificação antiga
        mask = (df_mapped['is_pdr'] == True) & (df_mapped['pdr_classification'] == old_classification)
        
        # Atualizar as mensagens
        df_mapped.loc[mask, 'message'] = df_mapped.loc[mask].apply(
            lambda row: update_message_classification(row['message'], old_classification, new_classification), 
            axis=1
        )
        
        # Atualizar também a classificação PDR extraída
        df_mapped.loc[mask, 'pdr_classification'] = new_classification
    
    return df_mapped

def connect_ssh_and_get_log(host, username, password, status_placeholder):
    """Conecta via SSH e executa o comando para gerar o log"""
    try:
        # Criar cliente SSH
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        
        # Conectar ao host
        client.connect(host, username=username, password=password)
        
        # Executar comando para gerar o log
        command = 'cd "$HOME/telas/Centro/" && mkdir -p "$HOME/Check_log_telas/" && cvs log -N -S > "$HOME/Check_log_telas/Check_log.csv"'
        
        # Executar comando e capturar saída em tempo real
        stdin, stdout, stderr = client.exec_command(command, get_pty=True)
        
        # Ler saída em tempo real
        output_lines = []
        while not stdout.channel.exit_status_ready():
            if stdout.channel.recv_ready():
                line = stdout.channel.recv(1024).decode('latin-1')
                output_lines.append(line)
                # Atualizar status com a última linha
                if line.strip():
                    status_placeholder.text(f"Executando: {line.strip()}")
        
        # Aguardar comando terminar completamente
        stdout.channel.recv_exit_status()
        
        # Ler o arquivo gerado com diferentes codificações
        command_cat = 'cat "$HOME/Check_log_telas/Check_log.csv"'
        stdin, stdout, stderr = client.exec_command(command_cat)
        
        # Tentar diferentes codificações
        encodings = ['latin-1', 'iso-8859-1', 'utf-8', 'cp1252']
        log_content = None
        
        for encoding in encodings:
            try:
                # Reset do buffer de saída
                stdin, stdout, stderr = client.exec_command(command_cat)
                log_content = stdout.read().decode(encoding)
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if log_content is None:
            # Última tentativa com errors='ignore'
            stdin, stdout, stderr = client.exec_command(command_cat)
            log_content = stdout.read().decode('latin-1', errors='ignore')
        
        # Fechar conexão
        client.close()
        
        return log_content
        
    except Exception as e:
        st.error(f"Erro na conexão SSH: {str(e)}")
        return None

def get_filtered_options(df, ignore_ana_dig=True, ignore_temp_files=True):
    """Retorna opções filtradas para os selectboxes baseado nos filtros aplicados"""
    filtered_df = df.copy()
    
    # Aplicar filtros básicos que afetam as opções
    if ignore_ana_dig:
        filtered_df = filtered_df.loc[~filtered_df['working_file'].str.startswith('Ana', na=False)]
        filtered_df = filtered_df.loc[~filtered_df['working_file'].str.startswith('Dig', na=False)]
    
    # Filtrar arquivos temporários (sempre True)
    if ignore_temp_files:
        filtered_df = filtered_df.loc[~filtered_df['working_file'].str.startswith('.#', na=False)]
        filtered_df = filtered_df.loc[~filtered_df['working_file'].str.startswith('.nfs', na=False)]
    
    # Extrair opções filtradas
    options = {
        'filenames': sorted(filtered_df['working_file'].unique()),
        'authors': sorted(filtered_df['author'].dropna().unique()),
        'centros': sorted(filtered_df['centro'].dropna().unique()) if 'centro' in filtered_df.columns else [],
        'estados': sorted(filtered_df['estado'].dropna().unique()) if 'estado' in filtered_df.columns else []
    }
    
    return options

def get_theme_adaptive_colors():
    """Retorna cores que funcionam bem em ambos os temas claro e escuro"""
    # Cores que funcionam bem em ambos os temas
    return {
        'text_color': "#AAAAAA",
        'grid_color': '#E0E0E0',
        'plot_bgcolor': 'rgba(0,0,0,0)',
        'paper_bgcolor': 'rgba(0,0,0,0)',
        'color_scale': 'sunsetdark',
        'qualitative_scale': 'Plotly'
    }

def main():
    st.set_page_config(page_title="Check Log de Telas", page_icon="📊", layout="wide")
    
    st.title("📊 Check Log de Telas")
    st.markdown("---")
    
    # Inicializar session state para armazenar dados
    if 'processed_data' not in st.session_state:
        st.session_state.processed_data = None
    if 'log_content' not in st.session_state:
        st.session_state.log_content = None
    if 'df' not in st.session_state:
        st.session_state.df = None
    if 'current_file_hash' not in st.session_state:
        st.session_state.current_file_hash = None
    if 'classification_mapping' not in st.session_state:
        st.session_state.classification_mapping = {}
    if 'show_classification_grouping' not in st.session_state:
        st.session_state.show_classification_grouping = False
    
    # Opção de carregamento do arquivo
    st.subheader("Carregamento do Arquivo de Log")
    
    option = st.radio(
        "Selecione o método de carregamento:",
        ["Gerar e carregar arquivo de log automaticamente", "Carregar arquivo de log manualmente"],
        horizontal=True
    )
    
    uploaded_file = None
    content = None
    new_file_detected = False
    
    if option == "Gerar e carregar arquivo de log automaticamente":
        col1, col2, col3 = st.columns(3)
        
        with col1:
            host = st.text_input("Host", value="rbsp01.reger.ons", disabled=True)
        with col2:
            user_id = st.text_input("User ID", placeholder="ex: REGER+ives.magalhaes")
        with col3:
            password = st.text_input("Senha", type="password")
        
        if st.button("Gerar e Carregar Log", help="O arquivo é gerado utilizando o comando _**cvs log**_ via CEUS e processado automaticamente"):
            if not user_id or not password:
                st.warning("Por favor, preencha User ID e Senha")
            else:
                status_placeholder = st.empty()
                with st.spinner("Conectando via SSH e gerando arquivo de log..."):
                    content = connect_ssh_and_get_log(host, user_id, password, status_placeholder)
                    status_placeholder.empty()  # Limpa o placeholder após conclusão
                    if content:
                        # Verificar se é um novo arquivo
                        content_hash = hash(content)
                        if st.session_state.current_file_hash != content_hash:
                            st.session_state.current_file_hash = content_hash
                            new_file_detected = True
                            
                        st.session_state.log_content = content
                        st.success("Log gerado e carregado com sucesso!")
    
    else:  # Carregar arquivo de log manualmente
        # Comando para o usuário copiar
        command = 'cd "$HOME/telas/Centro/" && mkdir -p "$HOME/Check_log_telas/" && cvs log -N -S > "$HOME/Check_log_telas/Check_log.csv" && /opt/Exceed_Connection_Server_13.8_64/bin/elpr $HOME/Check_log_telas/*'
        
        st.write("Copie o código abaixo e cole (Shift+Insert) no terminal do CEUS rbsp01.reger.ons")
        st.code(command, language="bash")
        
        uploaded_file = st.file_uploader("Carregar arquivo de log", type=['csv', 'txt'], key="file_uploader")
        
        # Botão para carregar novo arquivo (apenas no modo manual)
        if st.session_state.df is not None:
            if st.button("🔄 Carregar Novo Arquivo"):
                st.session_state.df = None
                st.session_state.log_content = None
                st.session_state.current_file_hash = None
                st.session_state.processed_data = None
                st.session_state.classification_mapping = {}
                st.session_state.show_classification_grouping = False
                st.rerun()
        
        if uploaded_file is not None:
            # Tentar diferentes codificações
            encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
            
            for encoding in encodings:
                try:
                    content = uploaded_file.read().decode(encoding)
                    uploaded_file.seek(0)  # Reset file pointer
                    
                    # Verificar se é um novo arquivo
                    content_hash = hash(content)
                    if st.session_state.current_file_hash != content_hash:
                        st.session_state.current_file_hash = content_hash
                        new_file_detected = True
                        
                    st.session_state.log_content = content
                    break
                except UnicodeDecodeError:
                    uploaded_file.seek(0)
                    continue
    
    # Usar dados da session state se disponíveis
    if st.session_state.log_content is not None:
        content = st.session_state.log_content
    
    # Processar conteúdo se disponível
    if content is not None:
        # Verificar se precisa reprocessar (novo arquivo ou dados não processados)
        if st.session_state.df is None or new_file_detected:
            with st.spinner('Processando arquivo...'):
                data = parse_log_content(content)
                df = pd.DataFrame(data)
                
                st.session_state.df = df
                st.session_state.processed_data = {
                    'data': data,
                    'df': df
                }
                st.success(f"Processados {len(df)} registros de revisão.")
        else:
            df = st.session_state.df
            st.info(f"Dados já processados anteriormente ({len(df)} registros)")
        
        # Botão para baixar o arquivo original
        if st.session_state.log_content:
            st.download_button(
                label="📥 Baixar Arquivo de Log Original",
                data=st.session_state.log_content,
                file_name=f"Log_telas_{datetime.now().strftime('%d_%m_%Y')}.csv",
                mime="text/csv"
            )
        
        # Filtros
        st.sidebar.header("Filtros")
        
        # Filtro Análise PDR
        pdr_only = st.sidebar.checkbox("Análise PDR", value=False, 
                                      help="Filtrar resultados com mensagem iniciando por '#' e realizar análise detalhada")
        
        # Filtro Ignorar Ana e Dig
        ignore_ana_dig = st.sidebar.checkbox("Ignorar Ana e Dig", value=True,
                                           help="Ignorar arquivos que começam com 'Ana' e 'Dig'")
        
        # Filtro Ignorar excluídos
        ignore_excluded = st.sidebar.checkbox("Ignorar excluídos", value=True,
                                            help="Ignorar arquivos excluídos (ficam registrados no diretório /Attic/)")
        
        # Ignorar temporários (fixo - sempre True)
        ignore_temp_files = True
               
        # Obter opções filtradas
        if not df.empty:
            filtered_options = get_filtered_options(df, ignore_ana_dig, ignore_temp_files)
            
            # Filtro por Centro
            if filtered_options['centros']:
                selected_centros = st.sidebar.multiselect(
                    "Centro",
                    options=filtered_options['centros'],
                    default=[],
                    placeholder="Selecione um ou mais centros",
                    help="Filtra pelo padrão seguinte a /telas/Centro/"
                )
            else:
                selected_centros = []
            
            # Filtro por Estado
            if filtered_options['estados']:
                selected_estados = st.sidebar.multiselect(
                    "Estado",
                    options=filtered_options['estados'],
                    default=[],
                    placeholder="Selecione um ou mais estados/subdiretórios",
                    help="Filtra pelo padrão seguinte a /telas/Centro/Nome_do_Centro/"
                )
            else:
                selected_estados = []
            
            # Filtro por nome do arquivo (multiseleção)
            if filtered_options['filenames']:
                selected_filenames = st.sidebar.multiselect(
                    "Nome da Tela",
                    options=filtered_options['filenames'],
                    default=[],
                    placeholder="Selecione um ou mais arquivos",
                    help="Filtra pelo nome do arquivo selecionado"
                )
            else:
                selected_filenames = []
            
            # Filtro por autor (multiseleção)
            if filtered_options['authors']:
                selected_authors = st.sidebar.multiselect(
                    "Autor",
                    options=filtered_options['authors'],
                    default=[],
                    placeholder="Selecione um ou mais autores",
                    help="Filtra pelo autor do commit"
                )
            else:
                selected_authors = []
            
            # Filtro por caminho
            path_filter = st.sidebar.text_input("Caminho da Tela", placeholder="Ex: /DRILL/", help="Filtrar pelo caminho do arquivo")

            # Filtro por data
            col1, col2 = st.sidebar.columns(2)
            with col1:
                if 'date' in df.columns and not df.empty:
                    # Converter datas para datetime para filtro
                    try:
                        df_date_filter = df.copy()
                        df_date_filter['date_dt'] = pd.to_datetime(df_date_filter['date'], format='%d/%m/%Y', errors='coerce')
                        min_date = df_date_filter['date_dt'].min()
                        max_date = df_date_filter['date_dt'].max()
                        
                        if pd.notna(min_date) and pd.notna(max_date):
                            default_end_date = max_date.date()
                            default_start_date = (max_date - timedelta(days=30)).date()
                            
                            start_date = st.date_input(
                                "Data Início",
                                value=default_start_date,
                                min_value=min_date.date(),
                                max_value=max_date.date(),
                                format="DD/MM/YYYY",
                                help="Por padrão é definido como 30 dias antes da data final encontrada."
                            )
                        else:
                            start_date = st.date_input("Data Início",format="DD/MM/YYYY")
                    except:
                        start_date = st.date_input("Data Início",format="DD/MM/YYYY")
                else:
                    start_date = st.date_input("Data Início",format="DD/MM/YYYY")
            
            with col2:
                if 'date' in df.columns and not df.empty:
                    try:
                        if pd.notna(min_date) and pd.notna(max_date):
                            end_date = st.date_input(
                                "Data Fim", 
                                value=default_end_date,
                                min_value=min_date.date(),
                                max_value=max_date.date(),
                                format="DD/MM/YYYY"
                            )
                        else:
                            end_date = st.date_input("Data Fim",format="DD/MM/YYYY")
                    except:
                        end_date = st.date_input("Data Fim",format="DD/MM/YYYY")
                else:
                    end_date = st.date_input("Data Fim",format="DD/MM/YYYY")
        
        # Aplicar filtros
        filtered_df = df.copy()
        
        if not df.empty:
            # Filtro PDR
            if pdr_only:
                filtered_df = filtered_df.loc[filtered_df['is_pdr'] == True]
            
            # Filtro Ignorar Ana e Dig
            if ignore_ana_dig:
                mask_ana = ~filtered_df['working_file'].str.startswith('Ana', na=False)
                mask_dig = ~filtered_df['working_file'].str.startswith('Dig', na=False)
                filtered_df = filtered_df.loc[mask_ana & mask_dig]
            
            # Filtro Ignorar arquivos temporários (sempre aplicado)
            if ignore_temp_files:
                mask_temp1 = ~filtered_df['working_file'].str.startswith('.#', na=False)
                mask_temp2 = ~filtered_df['working_file'].str.startswith('.nfs', na=False)
                filtered_df = filtered_df.loc[mask_temp1 & mask_temp2]
            
            # Filtro excluídos
            if ignore_excluded:
                mask_excluded = ~filtered_df['rcs_file'].str.contains('/Attic/', na=False)
                filtered_df = filtered_df.loc[mask_excluded]
            
            # Filtro por Centro
            if 'centro' in filtered_df.columns and selected_centros:
                filtered_df = filtered_df.loc[filtered_df['centro'].isin(selected_centros)]
            
            # Filtro por Estado
            if 'estado' in filtered_df.columns and selected_estados:
                filtered_df = filtered_df.loc[filtered_df['estado'].isin(selected_estados)]
            
            # Filtro por caminho
            if path_filter:
                mask_path = filtered_df['rcs_file'].str.contains(path_filter, case=False, na=False)
                filtered_df = filtered_df.loc[mask_path]
            
            # Filtro por nome do arquivo
            if selected_filenames:
                filtered_df = filtered_df.loc[filtered_df['working_file'].isin(selected_filenames)]
            
            # Filtro por autor
            if selected_authors:
                filtered_df = filtered_df.loc[filtered_df['author'].isin(selected_authors)]
            
            # Filtro por data
            if 'date' in filtered_df.columns:
                try:
                    filtered_df_date = filtered_df.copy()
                    filtered_df_date['date_dt'] = pd.to_datetime(filtered_df_date['date'], format='%d/%m/%Y', errors='coerce')
                    
                    if start_date:
                        start_datetime = datetime.combine(start_date, datetime.min.time())
                        mask_start = filtered_df_date['date_dt'] >= start_datetime
                        filtered_df = filtered_df.loc[mask_start]
                    
                    if end_date:
                        end_datetime = datetime.combine(end_date, datetime.max.time())
                        mask_end = filtered_df_date['date_dt'] <= end_datetime
                        filtered_df = filtered_df.loc[mask_end]
                except:
                    pass
            
            # Aplicar mapeamento de classificações se existir
            if st.session_state.classification_mapping:
                filtered_df = apply_classification_mapping_to_dataframe(filtered_df, st.session_state.classification_mapping)
            
            # Exibir resultados
            st.subheader(f"Resultados ({len(filtered_df)} registros)")

            # CORREÇÃO DA ORDENAÇÃO - FORMA MAIS ROBUSTA
            # Criar uma cópia para não modificar o original
            filtered_df_sorted = filtered_df.copy()
            
            # Converter a coluna 'date' para datetime
            filtered_df_sorted['date_dt'] = pd.to_datetime(
                filtered_df_sorted['date'], 
                format='%d/%m/%Y', 
                errors='coerce'
            )
            
            # Ordenar pela data convertida
            filtered_df_sorted = filtered_df_sorted.sort_values(by=['date_dt','time'], ascending=[False,False])
            
            # Remover a coluna temporária
            filtered_df_sorted = filtered_df_sorted.drop('date_dt', axis=1)
            
            # Criar DataFrame para exibição com o novo formato
            display_columns = ['centro', 'estado', 'rcs_file', 'working_file', 'revision', 'author', 'date', 'time']
            
            # Adicionar colunas PDR se for análise PDR
            if pdr_only:
                display_columns.extend(['pdr_classification', 'pdr_time', 'pdr_description'])
            else:
                display_columns.append('message')
            
            display_df = filtered_df_sorted[display_columns].copy()
            
            # Renomear colunas para exibição
            column_rename_map = {
                'centro': 'Centro',
                'estado': 'Estado', 
                'rcs_file': 'Caminho da Tela',
                'working_file': 'Nome da Tela',
                'revision': 'Revisão',
                'author': 'Autor',
                'date': 'Data',
                'time': 'Hora',
                'message': 'Mensagem',
                'pdr_classification': 'Tipo',
                'pdr_time': 'Tempo (min)',
                'pdr_description': 'Comentário'
            }
            
            display_df = display_df.rename(columns=column_rename_map)
            
            # Configurar a exibição do DataFrame (sem índice e ocultando Caminho da Tela)
            column_config = {
                "Centro": st.column_config.TextColumn(width="small"),
                "Estado": st.column_config.TextColumn(width="small"),
                "Caminho da Tela": st.column_config.Column(disabled=True),  # Oculta por padrão
                "Nome da Tela": st.column_config.TextColumn(width="medium"),
                "Revisão": st.column_config.TextColumn(width="none"),
                "Autor": st.column_config.TextColumn(width="medium"),
                "Data": st.column_config.TextColumn(width="none"),
                "Hora": st.column_config.TextColumn(width="none")
            }
            
            # Adicionar configurações para colunas PDR se for análise PDR
            if pdr_only:
                column_config.update({
                    "Tipo": st.column_config.TextColumn(width="none"),
                    "Tempo (min)": st.column_config.NumberColumn(width="small"),
                    "Comentário": st.column_config.TextColumn(width="medium")
                })
            else:
                column_config["Mensagem"] = st.column_config.TextColumn(width="large")
            
            # Ordem das colunas para exibição (sem Caminho da Tela)
            column_order = ['Centro', 'Estado', 'Nome da Tela', 'Revisão', 'Autor', 'Data', 'Hora']
            if pdr_only:
                column_order.extend(['Tipo', 'Tempo (min)', 'Comentário'])
            else:
                column_order.append('Mensagem')
            
            st.dataframe(
                display_df,
                use_container_width=True,
                height=600,
                hide_index=True,
                column_config=column_config,
                column_order=column_order
            )
           
            # Botão de popover para Classificação de Commits
            col_info, col_classif = st.columns([1, 5])
            with col_info:
                st.caption("ℹ️ O CrossVC considera o fuso horário GMT+0 (+3h em relação ao horário local).")
            with col_classif:
                with st.popover("📋 Classificação de Commits", use_container_width=False):
                    st.markdown("""
                    FORMATO: __**#CLASSIFICAÇÃO#TEMPO#COMENTÁRIO**__  
                    Onde a classificação é uma das listadas abaixo, o tempo (em minutos) é um número inteiro e o comentário é o campo livre para explicação do motivo da revisão.  
                      
                    **1. ANOMALIA**  
                    Refere-se a erros ou incoerências identificados externamente, por exemplo, pelas Salas de Operação ou outras gerências (como RSO, PRI), que impactam o funcionamento ou a coerência da tela.  
                    • Correções internamente identificadas devem ser classificadas como MANUTENÇÃO.  
                    • O tempo de execução da anomalia deve considerar todas as etapas envolvidas, como abertura do chamado no sistema OTRS, análise de diagrama envolvido, edição da tela, entre outras.  
                    • Quando houver mais de um tipo de alteração (ex: anomalia e melhoria), devem ser realizados commits separados, salvo quando uma das ações for irrelevante frente à outra.

                    **2. MANUT (Manutenção)**  
                    Refere-se a ajustes de rotina, preventivos ou corretivos, realizados pela equipe da PDR sem demanda externa.  
                    **Exemplos:**  
                    • Retirada de sinalização de futuro  
                    • Ajustes de textos de revisões de IO  
                    • Troca de agente operador  
                    • Troca de posicionamento de equipamentos  
                    • Correções de erros identificados internamente, desde que não tenham sido sinalizados por outras áreas

                    **3. RECOMP (Recomposição)**  
                    Classificação destinada a alterações em telas relacionadas ao processo de recomposição do sistema, como os corredores de recomposição.  
                    • **Exclusão**: casos em que houver erro identificado externamente (ANOMALIA), mesmo em telas de recomposição, devem ser registrados como ANOMALIA.

                    **4. NOVA**  
                    Aplica-se à criação de novas telas no REGER e equipamentos novos em telas já existentes.

                    **5. MELHORIA**  
                    Refere-se a alterações não essenciais, que não tratam erros, mas têm o objetivo de otimizar a usabilidade, visualização ou interpretação da tela.  
                    **Exemplos:**  
                    • Mudanças de layout  
                    • Inclusão de novos filtros  
                    • Inserção de elementos visuais ou alarmes  
                    • Ajustes de lógica solicitados pela operação, sem envolvimento de falha
                    """)

            # Botão de download em Excel
            today = datetime.now().strftime("%d_%m_%Y")
            filename = f"Check_log_telas-{today}.xlsx"

            
            # Botão de download em Excel
            today = datetime.now().strftime("%d_%m_%Y")
            filename = f"Check_log_telas-{today}.xlsx"
            
            # Criar arquivo Excel
            excel_file = create_excel_file(display_df)
            
            # Salvar em buffer
            from io import BytesIO
            excel_buffer = BytesIO()
            excel_file.save(excel_buffer)
            excel_buffer.seek(0)
            
            st.download_button(
                label="📥 Baixar em Excel",
                data=excel_buffer,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Botão de download em Excel
            today = datetime.now().strftime("%d_%m_%Y")
            filename = f"Check_log_telas-{today}.xlsx"
            
            # Análise PDR - Estatísticas detalhadas
            if pdr_only and len(filtered_df) > 0:
                st.subheader("📈 Análise PDR Detalhada")
                
                # Criar DataFrame apenas com registros PDR válidos (com classificação e tempo)
                pdr_df = filtered_df[
                    (filtered_df['pdr_classification'].notna()) & 
                    (filtered_df['pdr_time'].notna())
                ].copy()
                
                if len(pdr_df) > 0:
                    # Obter classificações únicas
                    classification_counts = pdr_df['pdr_classification'].value_counts()
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.write("**Classificações quantificadas:**")
                        # Formatar: MANUT: 10 | NOVA: 71 | MELHORIA: 25
                        classifications_text = " | ".join([f"{cls}: {count}" for cls, count in classification_counts.items()])
                        st.write(classifications_text)
                        
                        # Botão para agrupar classificações
                        if st.button("Agrupar Classificações"):
                            st.session_state.show_classification_grouping = not st.session_state.show_classification_grouping
                    
                    with col2:
                        # Arquivos mais modificados
                        file_counts = pdr_df['working_file'].value_counts().head(5)
                        st.write("**Arquivos Mais Modificados:**")
                        # Formatar: arquivo1: 19 | arquivo2: 20 | arquivo3: 8
                        files_text = " | ".join([f"{file}: {count}" for file, count in file_counts.items()])
                        st.write(files_text)
                    
                    # Seção de agrupamento de classificações (apenas se o botão foi clicado)
                    if st.session_state.show_classification_grouping:
                        st.subheader("🔄 Agrupamento de Classificações")
                        
                        # Obter classificações únicas atualizadas
                        current_unique = sorted(pdr_df['pdr_classification'].unique())
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            # Lista de seleção única para classificação original (destino)
                            target_classification = st.selectbox(
                                "Classificação original:",
                                options=current_unique,
                                key="target_classification"
                            )
                        
                        with col2:
                            # Lista de seleção múltipla para classificações a agrupar
                            # Remover a classificação destino das opções
                            source_options = [cls for cls in current_unique if cls != target_classification]
                            source_classifications = st.multiselect(
                                "Agrupar com:",
                                options=source_options,
                                key="source_classifications"
                            )
                        
                        # Botões lado a lado
                        col_btn1, col_btn2 = st.columns(2)
                        
                        with col_btn1:
                            # Botão para confirmar agrupamento
                            if st.button("Confirmar Agrupamento"):
                                if target_classification and source_classifications:
                                    # Adicionar mapeamento ao session state
                                    for source in source_classifications:
                                        st.session_state.classification_mapping[source] = target_classification
                                    
                                    st.success(f"Classificações {source_classifications} agrupadas em {target_classification}")
                                    
                                    # Forçar rerun para atualizar as listas
                                    st.rerun()
                        
                        with col_btn2:
                            # Botão para limpar agrupamentos (ao lado do confirmar)
                            if st.session_state.classification_mapping:
                                if st.button("🗑️ Limpar Agrupamentos"):
                                    st.session_state.classification_mapping = {}
                                    st.success("Agrupamentos limpos!")
                                    st.rerun()
                    
                    # Análise de tempo por classificação
                    st.subheader("⏱️ Análise de Tempo")
                    
                    # Tempo total por classificação
                    time_by_classification = pdr_df.groupby('pdr_classification')['pdr_time'].sum().sort_values(ascending=False)
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # Gráfico de pizza - Tempo por classificação
                        if len(time_by_classification) > 0:
                            colors = get_theme_adaptive_colors()
                            fig_pie = px.pie(
                                values=time_by_classification.values,
                                names=time_by_classification.index,
                                title="Distribuição de Tempo por Classificação",
                                color_discrete_sequence=px.colors.qualitative.Set3
                            )
                            fig_pie.update_layout(
                                paper_bgcolor=colors['paper_bgcolor'],
                                plot_bgcolor=colors['plot_bgcolor'],
                                font=dict(color=colors['text_color'])
                            )
                            fig_pie.update_traces(textposition='inside', textinfo='percent+label')
                            st.plotly_chart(fig_pie, use_container_width=True)
                    
                    with col2:
                        # Gráfico de barras - Tempo por classificação com tempo nas anotações
                        if len(time_by_classification) > 0:
                            colors = get_theme_adaptive_colors()
                            fig_bar = px.bar(
                                x=time_by_classification.index,
                                y=time_by_classification.values,
                                title="Tempo Total por Classificação (minutos)",
                                labels={'x': 'Classificação', 'y': 'Tempo Total (min)'},
                                color=time_by_classification.values,
                                color_continuous_scale=colors['color_scale']
                            )
                            
                            fig_bar.update_layout(
                                paper_bgcolor=colors['paper_bgcolor'],
                                plot_bgcolor=colors['plot_bgcolor'],
                                font=dict(color=colors['text_color']),
                                xaxis=dict(gridcolor=colors['grid_color']),
                                yaxis=dict(gridcolor=colors['grid_color'])
                            )
                            
                            # Adicionar anotações com o tempo
                            for i, (classification, time_val) in enumerate(zip(time_by_classification.index, time_by_classification.values)):
                                fig_bar.add_annotation(
                                    x=classification,
                                    y=time_val,
                                    text=f"{time_val:.0f} min",
                                    showarrow=False,
                                    yshift=10,
                                    font=dict(color=colors['text_color'], size=12)
                                )
                            
                            st.plotly_chart(fig_bar, use_container_width=True)
                    
                    # Top arquivos por tempo gasto
                    st.subheader("📋 Arquivos que Demandaram Mais Tempo")
                    
                    time_by_file = pdr_df.groupby('working_file')['pdr_time'].sum().sort_values(ascending=False).head(10)
                    
                    if len(time_by_file) > 0:
                        colors = get_theme_adaptive_colors()
                        fig_files = px.bar(
                            x=time_by_file.values,
                            y=time_by_file.index,
                            orientation='h',
                            title="Top 10 Arquivos por Tempo Gasto (minutos)",
                            labels={'x': 'Tempo Total (min)', 'y': 'Arquivo'},
                            color=time_by_file.values,
                            color_continuous_scale=colors['color_scale']
                        )
                        fig_files.update_layout(
                            paper_bgcolor=colors['paper_bgcolor'],
                            plot_bgcolor=colors['plot_bgcolor'],
                            font=dict(color=colors['text_color'])
                        )
                        st.plotly_chart(fig_files, use_container_width=True)
                    
                    # Análise por centro
                    st.subheader("🏢 Análise por Centro")
                    
                    # Análise por centro (já extraído)
                    centro_analysis = pdr_df[pdr_df['centro'].notna()]
                    
                    if len(centro_analysis) > 0:
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            # Quantidade por centro - ordenar decrescente
                            count_by_centro = centro_analysis['centro'].value_counts().sort_values(ascending=False)
                            colors = get_theme_adaptive_colors()
                            fig_count = px.bar(
                                x=count_by_centro.index,
                                y=count_by_centro.values,
                                title="Quantidade de Arquivos por Centro",
                                labels={'x': 'Centro', 'y': 'Quantidade de Arquivos'},
                                color=count_by_centro.values,
                                color_continuous_scale='teal'
                            )
                            fig_count.update_layout(
                                paper_bgcolor=colors['paper_bgcolor'],
                                plot_bgcolor=colors['plot_bgcolor'],
                                font=dict(color=colors['text_color'])
                            )
                            st.plotly_chart(fig_count, use_container_width=True)
                        
                        with col2:
                            # Tempo por centro - ordenar decrescente
                            time_by_centro = centro_analysis.groupby('centro')['pdr_time'].sum().sort_values(ascending=False)
                            colors = get_theme_adaptive_colors()
                            fig_time = px.bar(
                                x=time_by_centro.index,
                                y=time_by_centro.values,
                                title="Tempo Total por Centro (minutos)",
                                labels={'x': 'Centro', 'y': 'Tempo Total (min)'},
                                color=time_by_centro.values,
                                color_continuous_scale='algae'
                            )
                            
                            fig_time.update_layout(
                                paper_bgcolor=colors['paper_bgcolor'],
                                plot_bgcolor=colors['plot_bgcolor'],
                                font=dict(color=colors['text_color'])
                            )
                            
                            # Adicionar anotações com o tempo
                            for i, (centro, time_val) in enumerate(zip(time_by_centro.index, time_by_centro.values)):
                                fig_time.add_annotation(
                                    x=centro,
                                    y=time_val,
                                    text=f"{time_val:.0f} min",
                                    showarrow=False,
                                    yshift=10,
                                    font=dict(color=colors['text_color'], size=12)
                                )
                            
                            st.plotly_chart(fig_time, use_container_width=True)
                        
                        # Métricas por centro
                        st.write("**Métricas Detalhadas por Centro:**")
                        
                        centro_stats = centro_analysis.groupby('centro').agg({
                            'pdr_time': ['sum', 'mean', 'max', 'count'],
                            'working_file': 'nunique'
                        }).round(2)
                        
                        centro_stats.columns = ['Tempo Total (min)', 'Tempo Médio (min)', 'Tempo Máximo (min)', 'Total de Revisões', 'Arquivos Únicos']
                        st.dataframe(centro_stats, use_container_width=True)
                    
                    # Análise por estado
                    st.subheader("🗺️ Análise por Estado")
                    
                    estado_analysis = pdr_df[pdr_df['estado'].notna()]
                    
                    if len(estado_analysis) > 0:
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            # Quantidade por estado - ordenar decrescente
                            count_by_estado = estado_analysis['estado'].value_counts().sort_values(ascending=False).head(10)
                            colors = get_theme_adaptive_colors()
                            fig_count_estado = px.bar(
                                x=count_by_estado.index,
                                y=count_by_estado.values,
                                title="Top 10 Estados por Quantidade de Arquivos",
                                labels={'x': 'Estado', 'y': 'Quantidade de Arquivos'},
                                color=count_by_estado.values,
                                color_continuous_scale='purp'
                            )
                            fig_count_estado.update_layout(
                                paper_bgcolor=colors['paper_bgcolor'],
                                plot_bgcolor=colors['plot_bgcolor'],
                                font=dict(color=colors['text_color'])
                            )
                            st.plotly_chart(fig_count_estado, use_container_width=True)
                        
                        with col2:
                            # Tempo por estado - ordenar decrescente
                            time_by_estado = estado_analysis.groupby('estado')['pdr_time'].sum().sort_values(ascending=False).head(10)
                            colors = get_theme_adaptive_colors()
                            fig_time_estado = px.bar(
                                x=time_by_estado.index,
                                y=time_by_estado.values,
                                title="Top 10 Estados por Tempo Total (minutos)",
                                labels={'x': 'Estado', 'y': 'Tempo Total (min)'},
                                color=time_by_estado.values,
                                color_continuous_scale='sunsetdark'
                            )
                            
                            fig_time_estado.update_layout(
                                paper_bgcolor=colors['paper_bgcolor'],
                                plot_bgcolor=colors['plot_bgcolor'],
                                font=dict(color=colors['text_color'])
                            )
                            
                            # Adicionar anotações com o tempo
                            for i, (estado, time_val) in enumerate(zip(time_by_estado.index, time_by_estado.values)):
                                fig_time_estado.add_annotation(
                                    x=estado,
                                    y=time_val,
                                    text=f"{time_val:.0f} min",
                                    showarrow=False,
                                    yshift=10,
                                    font=dict(color=colors['text_color'], size=12)
                                )
                            
                            st.plotly_chart(fig_time_estado, use_container_width=True)
                    
                    # Estatísticas gerais
                    st.subheader("📊 Estatísticas Gerais PDR")
                    
                    total_time = pdr_df['pdr_time'].sum()
                    total_time_hours = total_time / 60
                    avg_time = pdr_df['pdr_time'].mean()
                    max_time = pdr_df['pdr_time'].max()
                    total_files = pdr_df['working_file'].nunique()
                    total_revisions = len(pdr_df)
                    
                    col1, col2, col3, col4, col5 = st.columns(5)
                    
                    with col1:
                        st.metric("Tempo Total", f"{total_time:.0f} min ({total_time_hours:.1f} h)")
                    with col2:
                        st.metric("Tempo Médio por Revisão", f"{avg_time:.1f} min")
                    with col3:
                        st.metric("Tempo Máximo por Revisão", f"{max_time:.1f} min")
                    with col4:
                        st.metric("Arquivos Únicos", total_files)
                    with col5:
                        st.metric("Total de Revisões", total_revisions)
                
                else:
                    st.info("Nenhum registro PDR com informações de classificação e tempo encontrado.")
        else:
            st.info("Nenhum dado para exibir após aplicar os filtros.")
    
    else:
        if option == "Carregar arquivo de log manualmente" and uploaded_file is None:
            st.info("Aguardando upload do arquivo de log...")
        elif option == "Gerar e carregar arquivo de log automaticamente" and content is None:
            st.info("Aguardando geração do arquivo de log...")

if __name__ == "__main__":
    main()